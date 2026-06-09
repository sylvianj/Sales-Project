# products/excel_views.py
"""Excel import + template download for Products (used by the POS Inventory page)."""
import io

import openpyxl
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated

from .models import Product

# The standard import columns. Order here is the order written to the template.
COLUMNS = [
    'sku', 'name', 'category', 'brand', 'price', 'cost',
    'tax_rate', 'stock_qty', 'reorder_level', 'reorder_qty',
    'barcode', 'description',
]
REQUIRED = ['sku', 'name', 'price']

# Columns that must be parsed as numbers (with sensible fallbacks)
DECIMAL_FIELDS = {'price', 'cost', 'tax_rate'}
INT_FIELDS = {'stock_qty', 'reorder_level', 'reorder_qty'}


def _clean(value):
    return str(value).strip() if value is not None else ''


class ProductExcelImportView(APIView):
    """POST a .xlsx file -> create/update products keyed by SKU."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'message': 'No file uploaded. Send it as form field "file".'}, status=400)

        try:
            wb = openpyxl.load_workbook(upload, data_only=True)
        except Exception as exc:
            return Response({'message': f'Could not read Excel file: {exc}'}, status=400)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'message': 'The spreadsheet is empty.'}, status=400)

        # Map header names -> column index (case-insensitive)
        header = [_clean(h).lower() for h in rows[0]]
        col_index = {name: header.index(name) for name in COLUMNS if name in header}

        missing_required = [c for c in REQUIRED if c not in col_index]
        if missing_required:
            return Response({
                'message': f"Missing required column(s): {', '.join(missing_required)}. "
                           f"Download the template to see the expected format."
            }, status=400)

        created = updated = 0
        errors = []

        for line_no, raw in enumerate(rows[1:], start=2):  # start=2 -> first data row is line 2 in Excel
            def cell(name):
                idx = col_index.get(name)
                return raw[idx] if idx is not None and idx < len(raw) else None

            sku = _clean(cell('sku'))
            name = _clean(cell('name'))
            if not sku and not name:
                continue  # skip fully blank rows

            try:
                if not sku or not name:
                    raise ValueError('sku and name are required')

                defaults = {'name': name}
                for field in ['category', 'brand', 'barcode', 'description']:
                    if field in col_index:
                        defaults[field] = _clean(cell(field)) or None
                for field in DECIMAL_FIELDS:
                    if field in col_index and _clean(cell(field)):
                        defaults[field] = float(cell(field))
                for field in INT_FIELDS:
                    if field in col_index and _clean(cell(field)):
                        defaults[field] = int(float(cell(field)))

                if 'price' not in defaults:
                    raise ValueError('price is required')

                _, was_created = Product.objects.update_or_create(sku=sku, defaults=defaults)
                if was_created:
                    created += 1
                else:
                    updated += 1
            except Exception as exc:
                errors.append({'row': line_no, 'message': str(exc)})

        return Response({
            'created': created,
            'updated': updated,
            'errors': errors,
            'message': f'Imported {created + updated} product(s): {created} created, {updated} updated.'
                       + (f' {len(errors)} row(s) had errors.' if errors else ''),
        })


class ProductTemplateView(APIView):
    """GET -> download a .xlsx template with the standard headers + one example row."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Products'
        ws.append(COLUMNS)
        # one example row so users see the expected shape
        ws.append([
            'SKU001', 'Wireless Mouse', 'Accessories', 'Logitech', 1500, 900,
            16, 50, 5, 10, '1234567890123', 'Example product - delete this row',
        ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="product_import_template.xlsx"'
        return response
